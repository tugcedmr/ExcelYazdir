import os
import re
from openpyxl import Workbook,load_workbook
wb = load_workbook("deneme\MorEvimAktifScript.xlsx")
ws = wb.active
os.system("md morevim1")
os.chdir(r'C:\Users\LENOVO\Desktop\morevim1')
for dongu in range(2,ws.max_row+1):
    os.chdir(r'C:\Users\LENOVO\Desktop\morevim1')
    Gsutun=ws["G"+format(dongu)].value
    Isutun=ws["I"+format(dongu)].value
    klasorSorgusu=Gsutun.replace(" ", "")
    def DosyaYazdir():
        dosya = open(str(Gsutun.replace(" ", ""))+(str(re.sub(r"\s+$", "",Isutun, flags=re.UNICODE))+format(dongu-1))+'.js','w+',encoding="utf-8")
        satir="\n //"+str(ws["G"+format(dongu)].value)
        satir=satir+"\n //"+str(ws["I"+format(dongu)].value)
        satir=satir+"\n //"+str(ws["J"+format(dongu)].value)
        satir=satir+"\n"+str(ws["K"+format(dongu)].value)
        dosya.write(satir)
        dosya.close()
    if os.path.exists(klasorSorgusu)==True:
        os.chdir(r'C:\Users\LENOVO\Desktop\morevim1')
        os.getcwd()
        os.chdir(r'C:/Users/LENOVO/Desktop/morevim1/'+klasorSorgusu)
        DosyaYazdir()
    os.chdir(r'C:\Users\LENOVO\Desktop\morevim1')
    if os.path.exists(klasorSorgusu)==False:
        os.system("md "+Gsutun.replace(" ", ""))
        os.chdir(r'C:/Users/LENOVO/Desktop/morevim1/'+klasorSorgusu)
        DosyaYazdir()
